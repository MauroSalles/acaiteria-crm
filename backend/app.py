"""
CRM Simples - Açaiteria Combina Açaí
Aplicação Flask para gerenciamento de clientes e vendas

Desenvolvido por: Grupo 22 - Projeto Integrador UNIVESP
Data: 2026
"""

from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from flask_restx import Api, Namespace, Resource
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from datetime import datetime, timedelta, timezone
from decimal import Decimal
from functools import wraps
import csv
import io
import logging
import os
import hashlib
import secrets
from pydantic import BaseModel, EmailStr, ValidationError, field_validator

from .models import db, Usuario, Cliente, Produto, Venda, ItemVenda, Pagamento, ConsentimentoHistorico, LogAcao, TicketSuporte, MensagemTicket

# =============================================================================
# CONFIGURAÇÃO INICIAL
# =============================================================================

from dotenv import load_dotenv

load_dotenv()

# Resolver caminhos absolutos para templates e arquivos estáticos
basedir = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
template_dir = os.path.join(basedir, 'frontend')
static_dir = os.path.join(basedir, 'frontend', 'static')

app = Flask(__name__, 
            template_folder=template_dir,
            static_folder=static_dir,
            static_url_path='/static')

# Chave secreta para sessões (obrigatória em produção)
_secret = os.environ.get('SECRET_KEY', '')
if not _secret and os.environ.get('FLASK_ENV') == 'production':
    raise RuntimeError('SECRET_KEY obrigatória em produção!')
app.config['SECRET_KEY'] = _secret or 'dev-secret-key-change-in-production'

# Proteção de cookies de sessão
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
if os.environ.get('FLASK_ENV') == 'production':
    app.config['SESSION_COOKIE_SECURE'] = True

# Desabilitar Swagger em produção
_doc_path = '/api/docs' if os.environ.get('FLASK_ENV') != 'production' else False

api = Api(
    app,
    version='1.0',
    title='Acaiteria CRM API',
    description='API REST do CRM — validação Pydantic, conformidade LGPD, rate-limiting.',
    doc=_doc_path,
    prefix='/api'
)

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["300 per hour"],
    storage_uri="memory://"
)

health_ns = Namespace('health', description='Healthcheck e metadados da API', path='/')
api.add_namespace(health_ns)

# Configuração do banco de dados
# Railway fornece postgres:// mas SQLAlchemy 2.x exige postgresql://
database_url = os.environ.get('DATABASE_URL', 'sqlite:///acaiteria.db')
if database_url.startswith('postgres://'):
    database_url = database_url.replace('postgres://', 'postgresql://', 1)

app.config['SQLALCHEMY_DATABASE_URI'] = database_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ECHO'] = os.environ.get('FLASK_ENV') == 'development'

# Inicializar banco de dados
db.init_app(app)

# Criar tabelas automaticamente no primeiro request (necessário em produção/cloud)
with app.app_context():
    db.create_all()

    # Auto-migração: adicionar colunas novas em tabelas existentes (SQLite/Postgres)
    _migracoes = [
        ('produto', 'estoque_atual', 'INTEGER DEFAULT 0'),
        ('produto', 'estoque_minimo', 'INTEGER DEFAULT 0'),
        ('cliente', 'pontos_fidelidade', 'INTEGER DEFAULT 0'),
    ]
    with db.engine.connect() as conn:
        for tabela, coluna, tipo in _migracoes:
            try:
                conn.execute(db.text(
                    f'ALTER TABLE {tabela} ADD COLUMN {coluna} {tipo}'
                ))
                conn.commit()
            except Exception:
                conn.rollback()

# Configurar logging estruturado
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s %(name)s: %(message)s'
)
logger = logging.getLogger('acaiteria-crm')


def _erro_interno(e):
    """Loga exceção e retorna resposta segura (sem stack trace em produção)."""
    logger.exception('Erro interno: %s', e)
    if app.config.get('TESTING') or os.environ.get('FLASK_ENV') == 'development':
        return jsonify({'erro': str(e)}), 500
    return jsonify({'erro': 'Erro interno do servidor'}), 500


# =============================================================================
# SECURITY HEADERS
# =============================================================================

@app.after_request
def adicionar_headers_seguranca(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'SAMEORIGIN'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    response.headers['Permissions-Policy'] = 'geolocation=(), camera=(), microphone=()'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self'; "
        "frame-ancestors 'none'"
    )
    if os.environ.get('FLASK_ENV') == 'production':
        response.headers['Strict-Transport-Security'] = 'max-age=31536000; includeSubDomains'
    return response


@app.context_processor
def inject_user():
    """Injeta dados do usuário logado em todos os templates."""
    return {
        'usuario_nome': session.get('usuario_nome', ''),
        'papel': session.get('papel', ''),
    }


class ClienteCreateSchema(BaseModel):
    nome: str
    telefone: str | None = None
    email: EmailStr | None = None
    observacoes: str | None = None
    consentimento_lgpd: bool = False
    versao_politica: str = 'v1.0'

    @field_validator('nome')
    @classmethod
    def nome_obrigatorio(cls, value: str) -> str:
        value = value.strip()
        if len(value) < 2:
            raise ValueError('Nome deve ter ao menos 2 caracteres')
        return value


class ProdutoCreateSchema(BaseModel):
    nome_produto: str
    categoria: str | None = None
    descricao: str | None = None
    preco: float
    estoque_atual: int | None = 0
    estoque_minimo: int | None = 0


class VendaItemSchema(BaseModel):
    id_produto: int
    quantidade: int


class VendaCreateSchema(BaseModel):
    id_cliente: int
    forma_pagamento: str = 'Dinheiro'
    observacoes: str | None = None
    desconto_percentual: float = 0.0
    taxa: float = 0.0
    itens: list[VendaItemSchema]

    @field_validator('desconto_percentual')
    @classmethod
    def desconto_valido(cls, v: float) -> float:
        if v < 0 or v > 100:
            raise ValueError('Desconto deve estar entre 0 e 100%')
        return v

    @field_validator('taxa')
    @classmethod
    def taxa_valida(cls, v: float) -> float:
        if v < 0:
            raise ValueError('Taxa não pode ser negativa')
        return v


class ConsentimentoSchema(BaseModel):
    consentimento_lgpd: bool
    versao_politica: str = 'v1.0'


class ProdutoUpdateSchema(BaseModel):
    nome_produto: str | None = None
    categoria: str | None = None
    descricao: str | None = None
    preco: float | None = None
    ativo: bool | None = None
    estoque_atual: int | None = None
    estoque_minimo: int | None = None


def validar_payload(schema_cls):
    dados = request.get_json(silent=True) or {}
    try:
        return schema_cls.model_validate(dados).model_dump()
    except ValidationError as e:
        raise ValueError(e.errors())


# =============================================================================
# AUTENTICAÇÃO — Login com email + senha e papéis (admin / operador)
# =============================================================================


def login_required(f):
    """Decorator que protege rotas HTML exigindo login."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('usuario_id'):
            return redirect('/login')
        return f(*args, **kwargs)
    return decorated


def api_login_required(f):
    """Decorator que protege rotas API exigindo login."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('usuario_id'):
            return jsonify({'erro': 'Autenticação necessária'}), 401
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    """Decorator que exige papel 'admin' para rotas HTML."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('usuario_id'):
            return redirect('/login')
        if session.get('papel') != 'admin':
            return render_template('index.html', stats=_stats_default(),
                                   error='Acesso restrito a administradores'), 403
        return f(*args, **kwargs)
    return decorated


def api_admin_required(f):
    """Decorator que exige papel 'admin' para rotas API."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('usuario_id'):
            return jsonify({'erro': 'Autenticação necessária'}), 401
        if session.get('papel') != 'admin':
            return jsonify({'erro': 'Acesso restrito a administradores'}), 403
        return f(*args, **kwargs)
    return decorated


def _stats_default():
    return {
        'total_clientes': 0, 'total_vendas': 0, 'faturamento_total': 0,
        'vendas_semana': 0, 'clientes_consentimento': 0, 'taxa_consentimento': 0,
    }


def registrar_log(acao, entidade, id_entidade=None, detalhes=None):
    """Registra uma ação no audit log (sem commit separado para evitar conflito)."""
    try:
        log = LogAcao(
            id_usuario=session.get('usuario_id'),
            acao=acao,
            entidade=entidade,
            id_entidade=id_entidade,
            detalhes=detalhes,
            ip=request.remote_addr,
        )
        db.session.add(log)
        db.session.commit()
    except Exception as e:
        logger.warning('Falha ao registrar log: %s', e)
        try:
            db.session.rollback()
        except Exception:
            pass


@health_ns.route('/health')
class HealthResource(Resource):
    def get(self):
        return {
            'status': 'ok',
            'service': 'acaiteria-crm',
            'timestamp': datetime.now(timezone.utc).isoformat()
        }, 200

# =============================================================================
# ROTAS - AUTENTICAÇÃO
# =============================================================================

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute", methods=["POST"])
def login():
    """Página de login com email e senha"""
    if request.method == 'POST':
        email = request.form.get('email', '').strip().lower()
        senha = request.form.get('senha', '')
        usuario = Usuario.query.filter_by(email=email, ativo=True).first()
        if usuario and usuario.verificar_senha(senha):
            session['usuario_id'] = usuario.id_usuario
            session['usuario_nome'] = usuario.nome
            session['papel'] = usuario.papel
            session['autenticado'] = True  # compatibilidade
            registrar_log('login', 'usuario', usuario.id_usuario, f'Login de {usuario.nome}')
            return redirect('/')
        logger.warning('Tentativa de login falha para %s de %s', email, request.remote_addr)
        return render_template('login.html', erro='Email ou senha incorretos.')
    return render_template('login.html')


@app.route('/logout')
def logout():
    """Encerrar sessão"""
    registrar_log('logout', 'usuario', session.get('usuario_id'), 'Logout')
    session.clear()
    return redirect('/login')


# =============================================================================
# ROTAS - PÁGINA INICIAL
# =============================================================================

@app.route('/')
@login_required
def index():
    """Página inicial - Dashboard"""
    try:
        # Estatísticas gerais
        total_clientes = Cliente.query.filter_by(ativo=True).count()
        total_vendas = Venda.query.count()
        faturamento_total = db.session.query(db.func.sum(Venda.valor_total)).scalar() or 0
        
        # Últimas vendas (últimos 7 dias)
        vendas_semana = Venda.query.filter(
            Venda.data_venda >= datetime.now(timezone.utc) - timedelta(days=7)
        ).count()
        
        # Clientes com permissão LGPD
        clientes_consentimento = Cliente.query.filter_by(
            ativo=True, 
            consentimento_lgpd=True
        ).count()
        
        stats = {
            'total_clientes': total_clientes,
            'total_vendas': total_vendas,
            'faturamento_total': float(faturamento_total) if faturamento_total else 0,
            'vendas_semana': vendas_semana,
            'clientes_consentimento': clientes_consentimento,
            'taxa_consentimento': round((clientes_consentimento / total_clientes * 100) if total_clientes > 0 else 0, 2)
        }
        
        return render_template('index.html', stats=stats,
                               usuario_nome=session.get('usuario_nome', ''),
                               papel=session.get('papel', ''))
    except Exception as e:
        logger.exception('Erro no dashboard: %s', e)
        return render_template('index.html', stats=_stats_default(), error=str(e),
                               usuario_nome=session.get('usuario_nome', ''),
                               papel=session.get('papel', ''))


# =============================================================================
# ROTAS - CLIENTES
# =============================================================================

@app.route('/api/clientes', methods=['GET'])
@limiter.limit("120 per minute")
@api_login_required
def listar_clientes():
    """Listar todos os clientes ativos com busca e paginação"""
    try:
        query = Cliente.query.filter_by(ativo=True)

        # Busca por nome, telefone ou email
        busca = request.args.get('busca', '').strip()
        if busca:
            filtro = f'%{busca}%'
            query = query.filter(
                db.or_(
                    Cliente.nome.ilike(filtro),
                    Cliente.telefone.ilike(filtro),
                    Cliente.email.ilike(filtro)
                )
            )

        # Paginação
        pagina = request.args.get('pagina', 1, type=int)
        por_pagina = request.args.get('por_pagina', 50, type=int)
        por_pagina = min(por_pagina, 100)  # limite máximo

        total = query.count()
        clientes = query.order_by(Cliente.nome).offset(
            (pagina - 1) * por_pagina
        ).limit(por_pagina).all()

        return jsonify({
            'clientes': [cliente.to_dict() for cliente in clientes],
            'total': total,
            'pagina': pagina,
            'por_pagina': por_pagina,
            'total_paginas': (total + por_pagina - 1) // por_pagina
        })
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/clientes', methods=['POST'])
@limiter.limit("30 per minute")
@api_login_required
def criar_cliente():
    """Criar novo cliente"""
    try:
        dados = validar_payload(ClienteCreateSchema)
        
        # Verificar LGPD
        consentimento = dados.get('consentimento_lgpd', False)
        data_consentimento = None
        
        if consentimento:
            data_consentimento = datetime.now(timezone.utc)
        
        # Criar cliente
        cliente = Cliente(
            nome=dados.get('nome'),
            telefone=dados.get('telefone'),
            email=dados.get('email'),
            observacoes=dados.get('observacoes'),
            consentimento_lgpd=consentimento,
            data_consentimento=data_consentimento,
            consentimento_versao=dados.get('versao_politica', 'v1.0') if consentimento else None,
            ativo=True
        )

        db.session.add(cliente)
        db.session.flush()  # gera id_cliente antes do commit

        # Registrar histórico LGPD quando há consentimento
        if consentimento:
            entrada = ConsentimentoHistorico(
                id_cliente=cliente.id_cliente,
                acao='concedeu',
                versao_politica=dados.get('versao_politica', 'v1.0'),
                ip_address=request.remote_addr,
                user_agent=request.headers.get('User-Agent', '')[:255]
            )
            db.session.add(entrada)

        db.session.commit()
        registrar_log('criar', 'cliente', cliente.id_cliente, f'Cliente criado: {cliente.nome}')

        return jsonify(cliente.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        if isinstance(e, ValueError):
            return jsonify({'erro': 'Payload invalido', 'detalhes': str(e)}), 400
        return _erro_interno(e)


@app.route('/api/clientes/<int:id_cliente>', methods=['GET'])
@api_login_required
def obter_cliente(id_cliente):
    """Obter detalhes de um cliente"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente:
            return jsonify({'erro': 'Cliente não encontrado'}), 404
        
        # Detalhes com histórico de vendas
        cliente_dict = cliente.to_dict()
        cliente_dict['total_vendas'] = len(cliente.vendas)
        cliente_dict['faturamento_total'] = float(
            sum(v.valor_total for v in cliente.vendas)
        )
        
        return jsonify(cliente_dict)
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/clientes/<int:id_cliente>', methods=['PUT'])
@api_login_required
def atualizar_cliente(id_cliente):
    """Atualizar dados de um cliente"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente:
            return jsonify({'erro': 'Cliente não encontrado'}), 404
        if not cliente.ativo:
            return jsonify({'erro': 'Cliente anonimizado não pode ser editado'}), 400
        
        dados = request.get_json(silent=True) or {}

        # Validar nome se fornecido
        if 'nome' in dados:
            nome = (dados['nome'] or '').strip()
            if len(nome) < 2:
                return jsonify({'erro': 'Nome deve ter ao menos 2 caracteres'}), 400
            cliente.nome = nome

        # Validar e-mail se fornecido
        if 'email' in dados:
            email = (dados['email'] or '').strip().lower() or None
            if email and '@' not in email:
                return jsonify({'erro': 'E-mail inválido'}), 400
            # Verificar duplicidade
            if email:
                existente = Cliente.query.filter(
                    Cliente.email == email, Cliente.ativo == True,
                    Cliente.id_cliente != id_cliente
                ).first()
                if existente:
                    return jsonify({'erro': 'E-mail já cadastrado por outro cliente'}), 409
            cliente.email = email

        # Validar telefone se fornecido
        if 'telefone' in dados:
            telefone = (dados['telefone'] or '').strip() or None
            if telefone:
                existente = Cliente.query.filter(
                    Cliente.telefone == telefone, Cliente.ativo == True,
                    Cliente.id_cliente != id_cliente
                ).first()
                if existente:
                    return jsonify({'erro': 'Telefone já cadastrado por outro cliente'}), 409
            cliente.telefone = telefone

        if 'observacoes' in dados:
            cliente.observacoes = dados.get('observacoes')
        
        db.session.commit()
        registrar_log('editar', 'cliente', id_cliente, f'Cliente editado: {cliente.nome}')
        
        return jsonify(cliente.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route('/api/clientes/<int:id_cliente>', methods=['DELETE'])
@api_login_required
def deletar_cliente(id_cliente):
    """Deletar (anonimizar) um cliente - LGPD"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente:
            return jsonify({'erro': 'Cliente não encontrado'}), 404
        
        # Anonimizar ao invés de deletar (LGPD)
        cliente.nome = f'CLIENTE_ANONIMIZADO_{id_cliente}'
        cliente.telefone = None
        cliente.email = None
        cliente.observacoes = None
        cliente.ativo = False
        cliente.data_exclusao = datetime.now(timezone.utc)
        
        db.session.commit()
        registrar_log('excluir', 'cliente', id_cliente, 'Cliente anonimizado (LGPD)')
        
        return jsonify({'mensagem': 'Cliente anonimizado conforme LGPD'}), 200
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# ROTAS - LGPD (consentimento e histórico de auditoria)
# =============================================================================

@app.route('/api/clientes/<int:id_cliente>/consentimento', methods=['PUT'])
@limiter.limit("30 per minute")
@api_login_required
def atualizar_consentimento(id_cliente):
    """Concede ou revoga consentimento LGPD e registra no histórico de auditoria"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente:
            return jsonify({'erro': 'Cliente não encontrado'}), 404

        dados = validar_payload(ConsentimentoSchema)
        consentiu = bool(dados.get('consentimento_lgpd'))
        versao = dados.get('versao_politica', 'v1.0')

        cliente.consentimento_lgpd = consentiu
        cliente.consentimento_versao = versao if consentiu else None
        cliente.data_consentimento = datetime.now(timezone.utc) if consentiu else None

        entrada = ConsentimentoHistorico(
            id_cliente=id_cliente,
            acao='concedeu' if consentiu else 'revogou',
            versao_politica=versao,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')[:255]
        )
        db.session.add(entrada)
        db.session.commit()

        return jsonify({
            'mensagem': f'Consentimento {"concedido" if consentiu else "revogado"} com sucesso',
            'cliente': cliente.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        if isinstance(e, ValueError):
            return jsonify({'erro': 'Payload invalido', 'detalhes': str(e)}), 400
        return _erro_interno(e)


@app.route('/api/clientes/<int:id_cliente>/consentimento/historico', methods=['GET'])
@limiter.limit("120 per minute")
@api_login_required
def historico_consentimento(id_cliente):
    """Retorna histórico completo de auditoria LGPD do cliente"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente:
            return jsonify({'erro': 'Cliente não encontrado'}), 404

        historico = ConsentimentoHistorico.query.filter_by(
            id_cliente=id_cliente
        ).order_by(ConsentimentoHistorico.data_acao.desc()).all()

        return jsonify({
            'id_cliente': id_cliente,
            'nome': cliente.nome,
            'consentimento_atual': cliente.consentimento_lgpd,
            'versao_atual': cliente.consentimento_versao,
            'historico': [h.to_dict() for h in historico]
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - PRODUTOS
# =============================================================================

@app.route('/api/produtos', methods=['GET'])
@limiter.limit("120 per minute")
@api_login_required
def listar_produtos():
    """Listar produtos com filtros opcionais: busca, categoria, incluir_inativos"""
    try:
        incluir_inativos = request.args.get('incluir_inativos', '').lower() == 'true'
        query = Produto.query if incluir_inativos else Produto.query.filter_by(ativo=True)

        busca = request.args.get('busca', '').strip()
        if busca:
            filtro = f'%{busca}%'
            query = query.filter(
                db.or_(
                    Produto.nome_produto.ilike(filtro),
                    Produto.descricao.ilike(filtro)
                )
            )

        categoria = request.args.get('categoria', '').strip()
        if categoria:
            query = query.filter(Produto.categoria.ilike(f'%{categoria}%'))

        produtos = query.order_by(Produto.nome_produto).all()
        return jsonify([produto.to_dict() for produto in produtos])
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/produtos/estoque-baixo', methods=['GET'])
@api_login_required
def produtos_estoque_baixo():
    """Lista produtos com estoque abaixo do mínimo (alertas)"""
    try:
        produtos = Produto.query.filter(
            Produto.ativo == True,
            Produto.estoque_atual <= Produto.estoque_minimo,
            (Produto.estoque_minimo > 0) | (Produto.estoque_atual > 0)
        ).order_by(Produto.estoque_atual).all()
        return jsonify([p.to_dict() for p in produtos])
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/produtos', methods=['POST'])
@limiter.limit("30 per minute")
@api_login_required
def criar_produto():
    """Criar novo produto"""
    try:
        dados = validar_payload(ProdutoCreateSchema)
        
        produto = Produto(
            nome_produto=dados.get('nome_produto'),
            categoria=dados.get('categoria'),
            descricao=dados.get('descricao'),
            preco=Decimal(str(dados.get('preco'))),
            estoque_atual=int(dados.get('estoque_atual', 0)) if dados.get('estoque_atual') is not None else 0,
            estoque_minimo=int(dados.get('estoque_minimo', 5)) if dados.get('estoque_minimo') is not None else 5,
            ativo=True
        )
        
        db.session.add(produto)
        db.session.commit()
        registrar_log('criar', 'produto', produto.id_produto, f'Produto criado: {produto.nome_produto}')
        
        return jsonify(produto.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        if isinstance(e, ValueError):
            return jsonify({'erro': 'Payload invalido', 'detalhes': str(e)}), 400
        return _erro_interno(e)


@app.route('/api/produtos/<int:id_produto>', methods=['GET'])
@api_login_required
def obter_produto(id_produto):
    """Obter detalhes de um produto"""
    try:
        produto = db.session.get(Produto, id_produto)
        if not produto:
            return jsonify({'erro': 'Produto não encontrado'}), 404
        return jsonify(produto.to_dict())
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/produtos/<int:id_produto>', methods=['PUT'])
@limiter.limit("30 per minute")
@api_login_required
def atualizar_produto(id_produto):
    """Atualizar dados de um produto"""
    try:
        produto = db.session.get(Produto, id_produto)
        if not produto:
            return jsonify({'erro': 'Produto não encontrado'}), 404

        dados = request.get_json(silent=True) or {}

        if 'nome_produto' in dados and dados['nome_produto']:
            produto.nome_produto = dados['nome_produto']
        if 'categoria' in dados:
            produto.categoria = dados['categoria']
        if 'descricao' in dados:
            produto.descricao = dados['descricao']
        if 'preco' in dados and dados['preco'] is not None:
            produto.preco = Decimal(str(dados['preco']))
        if 'ativo' in dados:
            produto.ativo = bool(dados['ativo'])
        if 'estoque_atual' in dados and dados['estoque_atual'] is not None:
            produto.estoque_atual = max(0, int(dados['estoque_atual']))
        if 'estoque_minimo' in dados and dados['estoque_minimo'] is not None:
            produto.estoque_minimo = max(0, int(dados['estoque_minimo']))

        db.session.commit()
        registrar_log('editar', 'produto', id_produto, f'Produto editado: {produto.nome_produto}')
        return jsonify(produto.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route('/api/produtos/<int:id_produto>', methods=['DELETE'])
@limiter.limit("30 per minute")
@api_login_required
def deletar_produto(id_produto):
    """Desativar produto (soft delete)"""
    try:
        produto = db.session.get(Produto, id_produto)
        if not produto:
            return jsonify({'erro': 'Produto não encontrado'}), 404

        produto.ativo = False
        db.session.commit()
        registrar_log('excluir', 'produto', id_produto, f'Produto desativado: {produto.nome_produto}')
        return jsonify({'mensagem': f'Produto desativado com sucesso'})
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# ROTAS - HISTÓRICO DE AÇÕES (AUDIT LOG)
# =============================================================================

@app.route('/api/logs', methods=['GET'])
@limiter.limit("60 per minute")
@api_admin_required
def listar_logs():
    """Listar histórico de ações (somente admin). Suporta paginação e filtros."""
    try:
        pagina = request.args.get('pagina', 1, type=int)
        por_pagina = min(request.args.get('por_pagina', 50, type=int), 100)
        entidade = request.args.get('entidade')
        acao = request.args.get('acao')

        query = LogAcao.query.order_by(LogAcao.data_hora.desc())

        if entidade:
            query = query.filter(LogAcao.entidade == entidade)
        if acao:
            query = query.filter(LogAcao.acao == acao)

        total = query.count()
        logs = query.offset((pagina - 1) * por_pagina).limit(por_pagina).all()

        return jsonify({
            'logs': [l.to_dict() for l in logs],
            'total': total,
            'pagina': pagina,
            'por_pagina': por_pagina,
            'total_paginas': (total + por_pagina - 1) // por_pagina,
        })
    except Exception as e:
        return _erro_interno(e)


@app.route('/historico')
@admin_required
def pagina_historico():
    """Página de histórico de ações (admin only)"""
    return render_template('historico.html')


# =============================================================================
# ROTAS - VENDAS
# =============================================================================

@app.route('/api/vendas', methods=['GET'])
@limiter.limit("120 per minute")
@api_login_required
def listar_vendas():
    """Listar vendas com filtros: data_inicio, data_fim, cliente, forma_pagamento, pagina"""
    try:
        query = Venda.query

        # Filtro por intervalo de datas
        data_inicio = request.args.get('data_inicio', '').strip()
        data_fim = request.args.get('data_fim', '').strip()
        if data_inicio:
            try:
                dt_ini = datetime.strptime(data_inicio, '%Y-%m-%d')
                query = query.filter(Venda.data_venda >= dt_ini)
            except ValueError:
                pass
        if data_fim:
            try:
                dt_fim = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
                query = query.filter(Venda.data_venda < dt_fim)
            except ValueError:
                pass

        # Filtro por cliente
        id_cliente = request.args.get('id_cliente', type=int)
        if id_cliente:
            query = query.filter_by(id_cliente=id_cliente)

        # Filtro por forma de pagamento
        forma = request.args.get('forma_pagamento', '').strip()
        if forma:
            query = query.filter(Venda.forma_pagamento.ilike(f'%{forma}%'))

        # Paginação
        pagina = request.args.get('pagina', 1, type=int)
        por_pagina = min(request.args.get('por_pagina', 50, type=int), 100)
        total = query.count()

        vendas = query.order_by(Venda.data_venda.desc()).offset(
            (pagina - 1) * por_pagina
        ).limit(por_pagina).all()

        return jsonify({
            'vendas': [venda.to_dict() for venda in vendas],
            'total': total,
            'pagina': pagina,
            'por_pagina': por_pagina,
            'total_paginas': (total + por_pagina - 1) // por_pagina
        })
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/vendas', methods=['POST'])
@limiter.limit("40 per minute")
@api_login_required
def criar_venda():
    """Criar nova venda"""
    try:
        dados = validar_payload(VendaCreateSchema)

        # Validação complementar de negócio
        if not dados.get('itens') or len(dados['itens']) == 0:
            return jsonify({'erro': 'Venda deve ter pelo menos um item'}), 400
        
        # Verificar se cliente existe
        cliente = db.session.get(Cliente, dados['id_cliente'])
        if not cliente:
            return jsonify({'erro': 'Cliente não encontrado'}), 404

        # LGPD: bloquear venda sem consentimento
        if not cliente.consentimento_lgpd:
            return jsonify({'erro': 'Venda não permitida: cliente sem consentimento LGPD'}), 400

        # Calcular valor total
        valor_total = Decimal('0.00')
        
        # Criar venda
        venda = Venda(
            id_cliente=dados['id_cliente'],
            forma_pagamento=dados.get('forma_pagamento', 'Dinheiro'),
            status_pagamento='Pendente',
            observacoes=dados.get('observacoes')
        )
        
        # Adicionar itens
        for item_dados in dados['itens']:
            produto = db.session.get(Produto, item_dados['id_produto'])
            if not produto:
                return jsonify({'erro': f'Produto {item_dados["id_produto"]} não encontrado'}), 404

            # Validar produto ativo
            if not produto.ativo:
                return jsonify({'erro': f'Produto "{produto.nome_produto}" está desativado e não pode ser vendido'}), 400
            
            quantidade = int(item_dados['quantidade'])

            # Verificar estoque somente se controle ativo (estoque ou mínimo > 0)
            controle_ativo = (produto.estoque_atual or 0) > 0 or (produto.estoque_minimo or 0) > 0
            if controle_ativo and produto.estoque_atual < quantidade:
                return jsonify({
                    'erro': f'Estoque insuficiente para "{produto.nome_produto}": disponível {produto.estoque_atual}, solicitado {quantidade}'
                }), 400

            preco_unitario = Decimal(str(produto.preco))
            subtotal = preco_unitario * quantidade
            
            item = ItemVenda(
                id_produto=item_dados['id_produto'],
                quantidade=quantidade,
                preco_unitario=preco_unitario,
                subtotal=subtotal
            )
            venda.itens.append(item)
            valor_total += subtotal

        # Aplicar desconto e taxa
        desconto_perc = Decimal(str(dados.get('desconto_percentual', 0)))
        taxa = Decimal(str(dados.get('taxa', 0)))
        desconto_valor = valor_total * desconto_perc / Decimal('100')
        valor_total = valor_total - desconto_valor + taxa

        venda.valor_total = valor_total
        
        # Criar pagamento
        pagamento = Pagamento(
            valor_pago=valor_total,
            metodo=dados.get('forma_pagamento', 'Dinheiro'),
            status='Concluído'
        )
        venda.pagamento = pagamento
        venda.status_pagamento = 'Concluído'
        
        # Descontar estoque dos produtos vendidos (somente se controlado)
        for item in venda.itens:
            prod = db.session.get(Produto, item.id_produto)
            if prod and ((prod.estoque_atual or 0) > 0 or (prod.estoque_minimo or 0) > 0):
                prod.estoque_atual = max(0, prod.estoque_atual - item.quantidade)

        # Acumular pontos de fidelidade (1 ponto por R$1 gasto)
        pontos_ganhos = int(valor_total)
        if pontos_ganhos > 0:
            cliente.pontos_fidelidade = (cliente.pontos_fidelidade or 0) + pontos_ganhos

        db.session.add(venda)
        db.session.commit()
        registrar_log('criar', 'venda', venda.id_venda, f'Venda #{venda.id_venda} - R${float(venda.valor_total):.2f}')
        
        resultado = venda.to_dict()
        resultado['pontos_ganhos'] = pontos_ganhos
        resultado['pontos_total'] = cliente.pontos_fidelidade or 0
        return jsonify(resultado), 201
    except Exception as e:
        db.session.rollback()
        if isinstance(e, ValueError):
            return jsonify({'erro': 'Payload invalido', 'detalhes': str(e)}), 400
        return _erro_interno(e)


@app.route('/api/vendas/<int:id_venda>', methods=['GET'])
@api_login_required
def obter_venda(id_venda):
    """Obter detalhes de uma venda"""
    try:
        venda = db.session.get(Venda, id_venda)
        if not venda:
            return jsonify({'erro': 'Venda não encontrada'}), 404
        
        return jsonify(venda.to_dict())
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/vendas/<int:id_venda>/cancelar', methods=['POST'])
@limiter.limit("10 per minute")
@api_admin_required
def cancelar_venda(id_venda):
    """Cancelar (estornar) uma venda — somente admin.
    Restaura estoque e remove pontos de fidelidade do cliente."""
    try:
        venda = db.session.get(Venda, id_venda)
        if not venda:
            return jsonify({'erro': 'Venda não encontrada'}), 404

        if venda.status_pagamento == 'Cancelado':
            return jsonify({'erro': 'Venda já foi cancelada anteriormente'}), 400

        dados = request.get_json(silent=True) or {}
        motivo = (dados.get('motivo') or '').strip()
        if not motivo or len(motivo) < 3:
            return jsonify({'erro': 'Motivo do cancelamento é obrigatório (mínimo 3 caracteres)'}), 400

        # Restaurar estoque dos itens
        for item in venda.itens:
            prod = db.session.get(Produto, item.id_produto)
            if prod:
                prod.estoque_atual = (prod.estoque_atual or 0) + item.quantidade

        # Remover pontos de fidelidade concedidos
        pontos_remover = int(venda.valor_total)
        cliente = db.session.get(Cliente, venda.id_cliente)
        if cliente and pontos_remover > 0:
            cliente.pontos_fidelidade = max(0, (cliente.pontos_fidelidade or 0) - pontos_remover)

        # Atualizar status da venda
        venda.status_pagamento = 'Cancelado'
        venda.observacoes = f'{venda.observacoes or ""}\n[CANCELADO] {motivo}'.strip()

        # Atualizar pagamento
        if venda.pagamento:
            venda.pagamento.status = 'Estornado'

        db.session.commit()
        registrar_log('cancelar', 'venda', id_venda,
                       f'Venda #{id_venda} cancelada — R${float(venda.valor_total):.2f} — Motivo: {motivo}')

        return jsonify({
            'mensagem': f'Venda #{id_venda} cancelada com sucesso',
            'estoque_restaurado': True,
            'pontos_removidos': pontos_remover,
            'venda': venda.to_dict()
        })
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# ROTAS - FIDELIDADE
# =============================================================================

@app.route('/api/clientes/<int:id_cliente>/pontos', methods=['GET'])
@api_login_required
def obter_pontos_fidelidade(id_cliente):
    """Consultar pontos de fidelidade de um cliente"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente or not cliente.ativo:
            return jsonify({'erro': 'Cliente não encontrado'}), 404
        return jsonify({
            'id_cliente': cliente.id_cliente,
            'nome': cliente.nome,
            'pontos': cliente.pontos_fidelidade or 0
        })
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/clientes/<int:id_cliente>/pontos/resgatar', methods=['POST'])
@limiter.limit("30 per minute")
@api_login_required
def resgatar_pontos(id_cliente):
    """Resgatar pontos de fidelidade (cada 100 pontos = R$5 de desconto)"""
    try:
        cliente = db.session.get(Cliente, id_cliente)
        if not cliente or not cliente.ativo:
            return jsonify({'erro': 'Cliente não encontrado'}), 404

        dados = request.get_json(silent=True) or {}
        pontos_resgatar = int(dados.get('pontos', 0))

        if pontos_resgatar <= 0:
            return jsonify({'erro': 'Quantidade de pontos deve ser positiva'}), 400

        pontos_disponiveis = cliente.pontos_fidelidade or 0
        if pontos_resgatar > pontos_disponiveis:
            return jsonify({'erro': f'Pontos insuficientes. Disponível: {pontos_disponiveis}'}), 400

        # Regra: cada 100 pontos = R$5.00 de desconto
        if pontos_resgatar < 100:
            return jsonify({'erro': 'Mínimo de 100 pontos para resgate'}), 400

        desconto = Decimal(str((pontos_resgatar // 100) * 5))
        pontos_usados = (pontos_resgatar // 100) * 100  # usa múltiplos de 100

        cliente.pontos_fidelidade = pontos_disponiveis - pontos_usados
        db.session.commit()
        registrar_log('resgatar', 'fidelidade', cliente.id_cliente,
                       f'{pontos_usados} pontos resgatados → R${float(desconto):.2f} desconto')

        return jsonify({
            'id_cliente': cliente.id_cliente,
            'pontos_resgatados': pontos_usados,
            'desconto_gerado': float(desconto),
            'pontos_restantes': cliente.pontos_fidelidade
        })
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route('/api/fidelidade/ranking', methods=['GET'])
@api_login_required
def ranking_fidelidade():
    """Top 10 clientes com mais pontos"""
    try:
        clientes = Cliente.query.filter(
            Cliente.ativo == True,
            Cliente.pontos_fidelidade > 0
        ).order_by(Cliente.pontos_fidelidade.desc()).limit(10).all()

        return jsonify([{
            'id_cliente': c.id_cliente,
            'nome': c.nome,
            'pontos': c.pontos_fidelidade or 0
        } for c in clientes])
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - BUSCA GLOBAL
# =============================================================================

@app.route('/api/busca', methods=['GET'])
@limiter.limit("60 per minute")
@api_login_required
def busca_global():
    """Busca global: pesquisa clientes e produtos simultaneamente"""
    try:
        q = request.args.get('q', '').strip()
        if len(q) < 2:
            return jsonify({'clientes': [], 'produtos': []})

        filtro = f'%{q}%'

        clientes = Cliente.query.filter(
            Cliente.ativo == True,
            db.or_(
                Cliente.nome.ilike(filtro),
                Cliente.telefone.ilike(filtro),
                Cliente.email.ilike(filtro)
            )
        ).order_by(Cliente.nome).limit(5).all()

        produtos = Produto.query.filter(
            Produto.ativo == True,
            db.or_(
                Produto.nome_produto.ilike(filtro),
                Produto.categoria.ilike(filtro)
            )
        ).order_by(Produto.nome_produto).limit(5).all()

        return jsonify({
            'clientes': [{'id': c.id_cliente, 'nome': c.nome, 'telefone': c.telefone, 'email': c.email} for c in clientes],
            'produtos': [{'id': p.id_produto, 'nome': p.nome_produto, 'preco': float(p.preco), 'categoria': p.categoria} for p in produtos],
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - DASHBOARD GRÁFICOS
# =============================================================================

@app.route('/api/dashboard/graficos', methods=['GET'])
@api_login_required
def dashboard_graficos():
    """Dados para gráficos do dashboard: vendas por dia (7d) e por forma de pagamento"""
    try:
        hoje = datetime.now(timezone.utc).date()
        inicio = hoje - timedelta(days=6)

        # Vendas por dia (últimos 7 dias)
        vendas_dia = db.session.query(
            db.func.date(Venda.data_venda).label('dia'),
            db.func.count(Venda.id_venda).label('qtd'),
            db.func.coalesce(db.func.sum(Venda.valor_total), 0).label('total')
        ).filter(
            db.func.date(Venda.data_venda) >= inicio
        ).group_by(
            db.func.date(Venda.data_venda)
        ).order_by(
            db.func.date(Venda.data_venda)
        ).all()

        # Montar dict dia → dados (preencher dias sem vendas com 0)
        mapa = {str(r.dia): {'qtd': r.qtd, 'total': float(r.total)} for r in vendas_dia}
        dias = []
        for i in range(7):
            d = inicio + timedelta(days=i)
            ds = str(d)
            dias.append({
                'data': ds,
                'label': d.strftime('%d/%m'),
                'quantidade': mapa.get(ds, {}).get('qtd', 0),
                'faturamento': mapa.get(ds, {}).get('total', 0),
            })

        # Vendas por forma de pagamento
        pagamentos = db.session.query(
            Pagamento.metodo,
            db.func.count(Pagamento.id_pagamento).label('qtd'),
            db.func.coalesce(db.func.sum(Pagamento.valor_pago), 0).label('total')
        ).group_by(Pagamento.metodo).all()

        por_pagamento = [
            {'forma': p.metodo, 'quantidade': p.qtd, 'total': float(p.total)}
            for p in pagamentos
        ]

        # Top 5 produtos mais vendidos
        top_produtos = db.session.query(
            Produto.nome_produto,
            db.func.sum(ItemVenda.quantidade).label('qtd')
        ).join(ItemVenda).filter(
            Produto.ativo == True
        ).group_by(Produto.nome_produto).order_by(
            db.func.sum(ItemVenda.quantidade).desc()
        ).limit(5).all()

        produtos_ranking = [
            {'produto': p.nome_produto, 'quantidade': int(p.qtd)}
            for p in top_produtos
        ]

        return jsonify({
            'vendas_por_dia': dias,
            'por_forma_pagamento': por_pagamento,
            'top_produtos': produtos_ranking,
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - RELATÓRIOS
# =============================================================================

@app.route('/api/relatorios/dia-atual', methods=['GET'])
@api_login_required
def relatorio_dia_atual():
    """Relatório de vendas do dia atual"""
    try:
        hoje = datetime.now(timezone.utc).date()
        
        vendas_hoje = Venda.query.filter(
            db.func.date(Venda.data_venda) == hoje
        ).all()
        
        total_vendas = len(vendas_hoje)
        faturamento = sum(v.valor_total for v in vendas_hoje)
        
        # Por forma de pagamento
        por_forma = {}
        for venda in vendas_hoje:
            forma = venda.forma_pagamento or 'Indefinido'
            if forma not in por_forma:
                por_forma[forma] = Decimal('0.00')
            por_forma[forma] += venda.valor_total
        
        return jsonify({
            'data': hoje.isoformat(),
            'total_vendas': total_vendas,
            'faturamento_total': float(faturamento),
            'por_forma_pagamento': {k: float(v) for k, v in por_forma.items()},
            'ticket_medio': float(faturamento / total_vendas) if total_vendas > 0 else 0
        })
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/relatorios/por-data', methods=['GET'])
@api_login_required
def relatorio_por_data():
    """Relatório de vendas filtrado por data (YYYY-MM-DD)"""
    try:
        data_str = request.args.get('data', '')
        if not data_str:
            return jsonify({'erro': 'Parâmetro "data" é obrigatório (formato YYYY-MM-DD)'}), 400

        from datetime import date as date_type
        try:
            data_filtro = date_type.fromisoformat(data_str)
        except ValueError:
            return jsonify({'erro': 'Formato de data inválido. Use YYYY-MM-DD'}), 400

        vendas_dia = Venda.query.filter(
            db.func.date(Venda.data_venda) == data_filtro
        ).all()

        total_vendas = len(vendas_dia)
        faturamento = sum(v.valor_total for v in vendas_dia)

        por_forma = {}
        for venda in vendas_dia:
            forma = venda.forma_pagamento or 'Indefinido'
            if forma not in por_forma:
                por_forma[forma] = {'quantidade': 0, 'total': Decimal('0.00')}
            por_forma[forma]['quantidade'] += 1
            por_forma[forma]['total'] += venda.valor_total

        return jsonify({
            'data': data_filtro.isoformat(),
            'total_vendas': total_vendas,
            'faturamento_total': float(faturamento),
            'por_forma_pagamento': {
                k: {'quantidade': v['quantidade'], 'total': float(v['total'])}
                for k, v in por_forma.items()
            },
            'ticket_medio': float(faturamento / total_vendas) if total_vendas > 0 else 0,
            'vendas': [v.to_dict() for v in vendas_dia]
        })
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/relatorios/clientes-frequentes', methods=['GET'])
@api_login_required
def relatorio_clientes_frequentes():
    """Clientes mais frequentes (últimos 30 dias) — com paginação"""
    try:
        dias = request.args.get('dias', 30, type=int)
        dias = min(max(dias, 1), 365)
        limite = request.args.get('limite', 10, type=int)
        limite = min(max(limite, 1), 100)
        pagina = request.args.get('pagina', 1, type=int)
        pagina = max(pagina, 1)

        data_limite = datetime.now(timezone.utc) - timedelta(days=dias)

        base_query = db.session.query(
            Cliente.id_cliente,
            Cliente.nome,
            Cliente.telefone,
            db.func.count(Venda.id_venda).label('total_compras'),
            db.func.sum(Venda.valor_total).label('faturamento'),
            db.func.max(Venda.data_venda).label('ultima_compra')
        ).join(Venda).filter(
            Venda.data_venda >= data_limite,
            Cliente.ativo == True
        ).group_by(
            Cliente.id_cliente,
            Cliente.nome,
            Cliente.telefone
        ).order_by(
            db.func.count(Venda.id_venda).desc()
        )

        total = base_query.count()
        clientes_freq = base_query.offset((pagina - 1) * limite).limit(limite).all()

        resultado = []
        for cliente in clientes_freq:
            resultado.append({
                'id_cliente': cliente.id_cliente,
                'nome': cliente.nome,
                'telefone': cliente.telefone,
                'total_compras': cliente.total_compras,
                'faturamento': float(cliente.faturamento),
                'ultima_compra': cliente.ultima_compra.isoformat() if cliente.ultima_compra else None
            })

        return jsonify({
            'dados': resultado,
            'total': total,
            'pagina': pagina,
            'limite': limite,
            'total_paginas': (total + limite - 1) // limite if total else 0,
        })
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/relatorios/produtos-ranking', methods=['GET'])
@api_login_required
def relatorio_produtos_ranking():
    """Produtos mais vendidos — com paginação"""
    try:
        limite = request.args.get('limite', 15, type=int)
        limite = min(max(limite, 1), 100)
        pagina = request.args.get('pagina', 1, type=int)
        pagina = max(pagina, 1)

        base_query = db.session.query(
            Produto.id_produto,
            Produto.nome_produto,
            db.func.count(ItemVenda.id_item).label('quantidade'),
            db.func.sum(ItemVenda.subtotal).label('faturamento')
        ).join(ItemVenda).filter(
            Produto.ativo == True
        ).group_by(
            Produto.id_produto,
            Produto.nome_produto
        ).order_by(
            db.func.count(ItemVenda.id_item).desc()
        )

        total = base_query.count()
        produtos_rank = base_query.offset((pagina - 1) * limite).limit(limite).all()

        resultado = []
        for produto in produtos_rank:
            resultado.append({
                'id_produto': produto.id_produto,
                'nome_produto': produto.nome_produto,
                'quantidade_vendida': produto.quantidade,
                'faturamento': float(produto.faturamento)
            })

        return jsonify({
            'dados': resultado,
            'total': total,
            'pagina': pagina,
            'limite': limite,
            'total_paginas': (total + limite - 1) // limite if total else 0,
        })
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - EXPORTAÇÃO
# =============================================================================

@app.route('/api/exportar/clientes-csv', methods=['GET'])
@api_login_required
def exportar_clientes_csv():
    """Exportar lista de clientes em CSV"""
    def sanitize_csv(value):
        """Previne CSV formula injection (=, +, -, @, tab, CR)"""
        if isinstance(value, str) and value and value[0] in ('=', '+', '-', '@', '\t', '\r'):
            return "'" + value
        return value

    try:
        clientes = Cliente.query.filter_by(ativo=True, consentimento_lgpd=True).all()
        
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Nome', 'Telefone', 'Email', 'Data de Cadastro'])
        
        for cliente in clientes:
            writer.writerow([
                sanitize_csv(cliente.nome),
                sanitize_csv(cliente.telefone or ''),
                sanitize_csv(cliente.email or ''),
                cliente.data_cadastro.strftime('%Y-%m-%d') if cliente.data_cadastro else ''
            ])
        
        output.seek(0)
        bytes_output = io.BytesIO(output.getvalue().encode('utf-8'))
        bytes_output.seek(0)
        
        return send_file(
            bytes_output,
            mimetype='text/csv',
            as_attachment=True,
            download_name=f'clientes_export_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")}.csv'
        )
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/exportar/relatorio-pdf', methods=['GET'])
@api_login_required
def exportar_relatorio_pdf():
    """Gera PDF com relatório de vendas do dia (ou data informada)"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm

    try:
        data_str = request.args.get('data', datetime.now(timezone.utc).strftime('%Y-%m-%d'))
        try:
            data_ref = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'erro': 'Data inválida. Use formato YYYY-MM-DD'}), 400

        # Buscar vendas do dia
        vendas = Venda.query.filter(
            db.func.date(Venda.data_venda) == data_ref
        ).order_by(Venda.data_venda).all()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                topMargin=1.5 * cm, bottomMargin=1.5 * cm)
        styles = getSampleStyleSheet()
        elements = []

        # Título
        elements.append(Paragraph(
            f'<b>Combina Açaí — Relatório de Vendas</b>', styles['Title']))
        elements.append(Paragraph(
            f'Data: {data_ref.strftime("%d/%m/%Y")}', styles['Normal']))
        elements.append(Spacer(1, 0.5 * cm))

        if vendas:
            # Tabela de vendas
            header = ['#', 'Cliente', 'Itens', 'Forma Pgto', 'Valor (R$)']
            rows = [header]
            total_geral = 0
            for v in vendas:
                cliente_nome = v.cliente.nome if v.cliente else '—'
                qtd_itens = sum(i.quantidade for i in v.itens) if v.itens else 0
                rows.append([
                    str(v.id_venda),
                    cliente_nome[:25],
                    str(qtd_itens),
                    v.forma_pagamento or '—',
                    f'{float(v.valor_total):.2f}',
                ])
                total_geral += float(v.valor_total)

            rows.append(['', '', '', 'TOTAL', f'{total_geral:.2f}'])

            t = Table(rows, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7B1FA2')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ALIGN', (0, 0), (0, -1), 'CENTER'),
                ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3E5F5')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ]))
            elements.append(t)
        else:
            elements.append(Paragraph('Nenhuma venda registrada nesta data.', styles['Normal']))

        elements.append(Spacer(1, 1 * cm))
        elements.append(Paragraph(
            f'Gerado em {datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")} — CRM Açaiteria',
            styles['Italic']))

        doc.build(elements)
        buf.seek(0)
        return send_file(
            buf,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'relatorio_{data_ref.strftime("%Y%m%d")}.pdf'
        )
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/exportar/clientes-xlsx', methods=['GET'])
@api_login_required
def exportar_clientes_xlsx():
    """Exporta clientes com consentimento LGPD em formato Excel (.xlsx)"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        clientes = Cliente.query.filter_by(ativo=True, consentimento_lgpd=True).order_by(Cliente.nome).all()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Clientes'

        # Header style
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='7B1FA2', end_color='7B1FA2', fill_type='solid')
        headers = ['ID', 'Nome', 'Telefone', 'Email', 'Pontos', 'Data Cadastro', 'Consentimento LGPD']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        for i, c in enumerate(clientes, 2):
            ws.cell(row=i, column=1, value=c.id_cliente)
            ws.cell(row=i, column=2, value=c.nome)
            ws.cell(row=i, column=3, value=c.telefone or '')
            ws.cell(row=i, column=4, value=c.email or '')
            ws.cell(row=i, column=5, value=getattr(c, 'pontos_fidelidade', 0) or 0)
            ws.cell(row=i, column=6, value=c.data_cadastro.strftime('%d/%m/%Y') if c.data_cadastro else '')
            ws.cell(row=i, column=7, value='Sim' if c.consentimento_lgpd else 'Não')

        # Auto-width columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or '')) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'clientes_{datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
    except ImportError:
        return jsonify({'erro': 'openpyxl não instalado. Execute: pip install openpyxl'}), 500
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# PÁGINAS HTML
# =============================================================================

@app.route('/cadastro-cliente')
@login_required
def pagina_cadastro_cliente():
    """Página de cadastro de cliente"""
    return render_template('cadastro_cliente.html')


@app.route('/nova-venda')
@login_required
def pagina_nova_venda():
    """Página de registro de venda"""
    return render_template('venda.html')


@app.route('/relatorios')
@login_required
def pagina_relatorios():
    """Página de relatórios"""
    return render_template('relatorios.html')


@app.route('/clientes')
@login_required
def pagina_clientes():
    """Página de gerenciamento de clientes"""
    return render_template('clientes.html')


@app.route('/produtos')
@login_required
def pagina_produtos():
    """Página de gerenciamento de produtos"""
    return render_template('produtos.html')


@app.route('/fechamento')
@login_required
def pagina_fechamento():
    """Página de fechamento diário"""
    return render_template('fechamento.html')


@app.route('/politica-privacidade')
def politica_privacidade():
    """Página com política de privacidade LGPD"""
    return render_template('politica_privacidade.html')


@app.route('/usuarios')
@admin_required
def pagina_usuarios():
    """Página de gerenciamento de usuários (admin only)"""
    return render_template('usuarios.html')


# =============================================================================
# TOTEM — Cadastro público de clientes (sem login)
# =============================================================================

@app.route('/totem')
def pagina_totem():
    """Totem de auto-cadastro — página pública para tablet no balcão"""
    return render_template('totem_cliente.html')


@app.route('/api/totem/cadastro', methods=['POST'])
@limiter.limit("10 per minute")
def totem_cadastrar_cliente():
    """API pública para auto-cadastro via totem (rate-limited)"""
    try:
        dados = request.get_json(silent=True) or {}

        nome = (dados.get('nome') or '').strip()
        if not nome or len(nome) < 2:
            return jsonify({'erro': 'Nome é obrigatório (mínimo 2 caracteres)'}), 400

        consentimento = dados.get('consentimento_lgpd', False)
        if not consentimento:
            return jsonify({'erro': 'Consentimento LGPD é obrigatório'}), 400

        telefone = (dados.get('telefone') or '').strip() or None
        email = (dados.get('email') or '').strip().lower() or None
        observacoes = (dados.get('observacoes') or '').strip() or None

        # Validar email se fornecido
        if email and '@' not in email:
            return jsonify({'erro': 'E-mail inválido'}), 400

        # Verificar duplicidade por telefone ou email
        if telefone:
            existente = Cliente.query.filter_by(telefone=telefone, ativo=True).first()
            if existente:
                return jsonify({'erro': 'Este telefone já está cadastrado. Fale com o atendente.'}), 409
        if email:
            existente = Cliente.query.filter_by(email=email, ativo=True).first()
            if existente:
                return jsonify({'erro': 'Este e-mail já está cadastrado. Fale com o atendente.'}), 409

        cliente = Cliente(
            nome=nome,
            telefone=telefone,
            email=email,
            observacoes=observacoes,
            consentimento_lgpd=True,
            data_consentimento=datetime.now(timezone.utc),
            consentimento_versao=dados.get('versao_politica', 'v1.0'),
            ativo=True
        )

        db.session.add(cliente)
        db.session.flush()

        # Registrar histórico LGPD
        entrada = ConsentimentoHistorico(
            id_cliente=cliente.id_cliente,
            acao='concedeu',
            versao_politica=dados.get('versao_politica', 'v1.0'),
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent', '')[:255]
        )
        db.session.add(entrada)

        # Bônus de boas-vindas: +10 pontos
        cliente.pontos_fidelidade = 10

        db.session.commit()
        registrar_log('criar', 'cliente', cliente.id_cliente,
                       f'Auto-cadastro via totem: {cliente.nome}')

        return jsonify({
            'id_cliente': cliente.id_cliente,
            'nome': cliente.nome,
            'pontos_fidelidade': cliente.pontos_fidelidade,
            'mensagem': 'Cadastro realizado com sucesso!'
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'erro': 'Erro interno. Tente novamente.'}), 500


# =============================================================================
# ROTAS - GESTÃO DE USUÁRIOS (admin only)
# =============================================================================

class UsuarioCreateSchema(BaseModel):
    nome: str
    email: EmailStr
    senha: str
    papel: str = 'operador'

    @field_validator('nome')
    @classmethod
    def nome_valido(cls, v: str) -> str:
        v = v.strip()
        if len(v) < 2:
            raise ValueError('Nome deve ter ao menos 2 caracteres')
        return v

    @field_validator('senha')
    @classmethod
    def senha_forte(cls, v: str) -> str:
        if len(v) < 4:
            raise ValueError('Senha deve ter ao menos 4 caracteres')
        return v

    @field_validator('papel')
    @classmethod
    def papel_valido(cls, v: str) -> str:
        if v not in ('admin', 'operador'):
            raise ValueError('Papel deve ser "admin" ou "operador"')
        return v


@app.route('/api/usuarios', methods=['GET'])
@api_admin_required
def listar_usuarios():
    """Listar todos os usuários (admin only)"""
    try:
        usuarios = Usuario.query.order_by(Usuario.nome).all()
        return jsonify([u.to_dict() for u in usuarios])
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/usuarios', methods=['POST'])
@limiter.limit("10 per minute")
@api_admin_required
def criar_usuario():
    """Criar novo usuário (admin only)"""
    try:
        dados = validar_payload(UsuarioCreateSchema)

        if Usuario.query.filter_by(email=dados['email'].lower()).first():
            return jsonify({'erro': 'Email já cadastrado'}), 409

        usuario = Usuario(
            nome=dados['nome'],
            email=dados['email'].lower(),
            papel=dados.get('papel', 'operador'),
        )
        usuario.set_senha(dados['senha'])

        db.session.add(usuario)
        db.session.commit()
        return jsonify(usuario.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        if isinstance(e, ValueError):
            return jsonify({'erro': 'Payload inválido', 'detalhes': str(e)}), 400
        return _erro_interno(e)


@app.route('/api/usuarios/<int:id_usuario>', methods=['PUT'])
@limiter.limit("10 per minute")
@api_admin_required
def atualizar_usuario(id_usuario):
    """Atualizar usuário (admin only)"""
    try:
        usuario = db.session.get(Usuario, id_usuario)
        if not usuario:
            return jsonify({'erro': 'Usuário não encontrado'}), 404

        dados = request.get_json(silent=True) or {}

        if 'nome' in dados and dados['nome']:
            usuario.nome = dados['nome'].strip()
        if 'email' in dados and dados['email']:
            novo_email = dados['email'].strip().lower()
            existente = Usuario.query.filter_by(email=novo_email).first()
            if existente and existente.id_usuario != id_usuario:
                return jsonify({'erro': 'Email já cadastrado por outro usuário'}), 409
            usuario.email = novo_email
        if 'papel' in dados and dados['papel'] in ('admin', 'operador'):
            usuario.papel = dados['papel']
        if 'senha' in dados and dados['senha']:
            if len(dados['senha']) < 4:
                return jsonify({'erro': 'Senha deve ter ao menos 4 caracteres'}), 400
            usuario.set_senha(dados['senha'])
        if 'ativo' in dados:
            usuario.ativo = bool(dados['ativo'])

        db.session.commit()
        return jsonify(usuario.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route('/api/usuarios/<int:id_usuario>', methods=['DELETE'])
@api_admin_required
def deletar_usuario(id_usuario):
    """Desativar usuário (admin only, não pode desativar a si mesmo)"""
    try:
        if session.get('usuario_id') == id_usuario:
            return jsonify({'erro': 'Não é possível desativar seu próprio usuário'}), 400

        usuario = db.session.get(Usuario, id_usuario)
        if not usuario:
            return jsonify({'erro': 'Usuário não encontrado'}), 404

        usuario.ativo = False
        db.session.commit()
        return jsonify({'mensagem': f'Usuário {usuario.nome} desativado'})
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route('/api/me', methods=['GET'])
@api_login_required
def usuario_atual():
    """Retorna dados do usuário logado"""
    try:
        usuario = db.session.get(Usuario, session.get('usuario_id'))
        if not usuario:
            return jsonify({'erro': 'Usuário não encontrado'}), 404
        return jsonify(usuario.to_dict())
    except Exception as e:
        return _erro_interno(e)


# =============================================================================
# ROTAS - SUPORTE (Tickets + Chat)
# =============================================================================

@app.route('/suporte')
@login_required
def pagina_suporte():
    return render_template('suporte.html')


class TicketCreateSchema(BaseModel):
    assunto: str
    categoria: str = 'duvida'
    prioridade: str = 'normal'
    mensagem: str

    @field_validator('assunto')
    @classmethod
    def assunto_valido(cls, v: str) -> str:
        v = v.strip()
        if len(v) < 3:
            raise ValueError('Assunto deve ter ao menos 3 caracteres')
        if len(v) > 200:
            raise ValueError('Assunto deve ter no máximo 200 caracteres')
        return v

    @field_validator('categoria')
    @classmethod
    def categoria_valida(cls, v: str) -> str:
        if v not in ('duvida', 'problema', 'sugestao', 'outro'):
            raise ValueError('Categoria inválida')
        return v

    @field_validator('prioridade')
    @classmethod
    def prioridade_valida(cls, v: str) -> str:
        if v not in ('baixa', 'normal', 'alta', 'urgente'):
            raise ValueError('Prioridade inválida')
        return v

    @field_validator('mensagem')
    @classmethod
    def mensagem_valida(cls, v: str) -> str:
        v = v.strip()
        if len(v) < 5:
            raise ValueError('Mensagem deve ter ao menos 5 caracteres')
        return v


@app.route('/api/suporte/tickets', methods=['GET'])
@api_login_required
def listar_tickets():
    """Lista tickets: admin vê todos, operador vê os próprios"""
    try:
        pagina = request.args.get('pagina', 1, type=int)
        limite = min(request.args.get('limite', 20, type=int), 100)
        status_filtro = request.args.get('status')

        query = TicketSuporte.query
        if session.get('papel') != 'admin':
            query = query.filter_by(id_usuario=session['usuario_id'])
        if status_filtro:
            query = query.filter_by(status=status_filtro)

        query = query.order_by(TicketSuporte.data_atualizacao.desc())
        total = query.count()
        tickets = query.offset((pagina - 1) * limite).limit(limite).all()

        return jsonify({
            'dados': [t.to_dict() for t in tickets],
            'total': total,
            'pagina': pagina,
            'limite': limite,
            'total_paginas': (total + limite - 1) // limite,
        })
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/suporte/tickets', methods=['POST'])
@limiter.limit("10 per minute")
@api_login_required
def criar_ticket():
    """Abre um novo ticket de suporte"""
    try:
        dados = validar_payload(TicketCreateSchema)

        ticket = TicketSuporte(
            id_usuario=session['usuario_id'],
            assunto=dados['assunto'],
            categoria=dados['categoria'],
            prioridade=dados['prioridade'],
        )
        db.session.add(ticket)
        db.session.flush()

        msg = MensagemTicket(
            id_ticket=ticket.id_ticket,
            id_usuario=session['usuario_id'],
            conteudo=dados['mensagem'],
        )
        db.session.add(msg)
        db.session.commit()

        registrar_log('criar', 'ticket_suporte', ticket.id_ticket, dados['assunto'])
        return jsonify(ticket.to_dict()), 201
    except ValueError as e:
        return jsonify({'erro': 'Dados inválidos', 'detalhes': str(e)}), 400
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route('/api/suporte/tickets/<int:id_ticket>', methods=['GET'])
@api_login_required
def obter_ticket(id_ticket):
    """Retorna ticket com mensagens (chat)"""
    try:
        ticket = db.session.get(TicketSuporte, id_ticket)
        if not ticket:
            return jsonify({'erro': 'Ticket não encontrado'}), 404

        if session.get('papel') != 'admin' and ticket.id_usuario != session.get('usuario_id'):
            return jsonify({'erro': 'Acesso negado'}), 403

        return jsonify(ticket.to_dict())
    except Exception as e:
        return _erro_interno(e)


@app.route('/api/suporte/tickets/<int:id_ticket>/mensagens', methods=['POST'])
@limiter.limit("30 per minute")
@api_login_required
def enviar_mensagem_ticket(id_ticket):
    """Envia mensagem em um ticket (chat)"""
    try:
        ticket = db.session.get(TicketSuporte, id_ticket)
        if not ticket:
            return jsonify({'erro': 'Ticket não encontrado'}), 404

        if session.get('papel') != 'admin' and ticket.id_usuario != session.get('usuario_id'):
            return jsonify({'erro': 'Acesso negado'}), 403

        if ticket.status == 'fechado':
            return jsonify({'erro': 'Este ticket está fechado'}), 400

        dados = request.get_json(silent=True) or {}
        conteudo = (dados.get('conteudo') or '').strip()
        if len(conteudo) < 1:
            return jsonify({'erro': 'Mensagem não pode ser vazia'}), 400

        msg = MensagemTicket(
            id_ticket=id_ticket,
            id_usuario=session['usuario_id'],
            conteudo=conteudo,
        )
        db.session.add(msg)

        if ticket.status == 'aberto' and session.get('papel') == 'admin':
            ticket.status = 'em_andamento'

        db.session.commit()
        return jsonify(msg.to_dict()), 201
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


@app.route('/api/suporte/tickets/<int:id_ticket>/status', methods=['PUT'])
@api_login_required
def atualizar_status_ticket(id_ticket):
    """Atualiza status do ticket (admin pode tudo, operador só pode fechar os próprios)"""
    try:
        ticket = db.session.get(TicketSuporte, id_ticket)
        if not ticket:
            return jsonify({'erro': 'Ticket não encontrado'}), 404

        dados = request.get_json(silent=True) or {}
        novo_status = dados.get('status')
        if novo_status not in ('aberto', 'em_andamento', 'resolvido', 'fechado'):
            return jsonify({'erro': 'Status inválido'}), 400

        if session.get('papel') != 'admin':
            if ticket.id_usuario != session.get('usuario_id'):
                return jsonify({'erro': 'Acesso negado'}), 403
            if novo_status not in ('fechado',):
                return jsonify({'erro': 'Operador só pode fechar o próprio ticket'}), 403

        ticket.status = novo_status
        db.session.commit()

        registrar_log('atualizar', 'ticket_suporte', id_ticket, f'status → {novo_status}')
        return jsonify(ticket.to_dict())
    except Exception as e:
        db.session.rollback()
        return _erro_interno(e)


# =============================================================================
# TRATAMENTO DE ERROS
# =============================================================================

@app.errorhandler(404)
def nao_encontrado(erro):
    if request.path.startswith('/api/'):
        return jsonify({'erro': 'Endpoint não encontrado'}), 404
    return render_template('404.html'), 404


@app.errorhandler(500)
def erro_interno(erro):
    if request.path.startswith('/api/'):
        return jsonify({'erro': 'Erro interno do servidor'}), 500
    return render_template('500.html'), 500


# =============================================================================
# SEED — Cria admin padrão se não existir nenhum usuário
# =============================================================================

def _seed_admin():
    """Cria usuário admin padrão caso a tabela esteja vazia."""
    if Usuario.query.first() is None:
        admin = Usuario(
            nome='Administrador',
            email=os.environ.get('ADMIN_EMAIL', 'admin@acaiteria.com'),
            papel='admin',
        )
        admin.set_senha(os.environ.get('ADMIN_SENHA', 'admin123'))
        db.session.add(admin)
        db.session.commit()
        logger.info('Admin padrão criado: %s', admin.email)


# =============================================================================
# CRIAR TABELAS E SEED (usado pelo gunicorn na nuvem — tabelas já criadas acima)
# =============================================================================

with app.app_context():
    _seed_admin()


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
